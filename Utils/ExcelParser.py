from openpyxl import *
from openpyxl.styles import PatternFill,Font,Border,Side,Alignment,colors
from Utils.DateAndTime import *
import os




class ExcelParser():

    def __init__(self,filePath):
        self.filePath=filePath
        if os.path.exists(filePath) and ("xlsx" in filePath):
            self.wb = load_workbook(filePath)
            print("加载excel文件【%s】成功"%(os.path.basename(filePath)))
            self.sheet = self.wb.active
        else:
            print("指定的excel文件【%s】不存在，或者文件类型不是excel" % filePath)
        
    def getSheetNames(self):
        return self.wb.sheetnames

    def getSheetByIndex(self,index):
        if not isinstance(index,int):
            print("sheet序号【%s】不是整数，请重试！"%index)
            return
        if index>(len(self.wb.sheetnames)) or index<0:
            print("输入的序号超出范围，请重试！")
            return
        else:
            self.sheet=self.wb[self.getSheetNames()[index-1]]
        return self.sheet

    def getSheetByName(self,sheetName):
        if not sheetName in self.getSheetNames():
            print("设定的sheetName【%s】不存在！"%sheetName)
            return
        self.sheet=self.wb[sheetName]
        return self.sheet

    def getMaxRowNo(self):
        return self.sheet.max_row

    def getMaxColumnNo(self):
        return self.sheet.max_column

    def getSheetAllCellValuesByRow(self):
        data=[]
        for row in self.sheet.iter_rows():
            row_data=[]
            for cell in row:
                row_data.append(cell.value)
            data.append(row_data)
        return data

    def getSheetAllCellValuesByColumn(self):
        data=[]
        for col in self.sheet.iter_cols():
            col_data=[]
            for cell in col:
                col_data.append(cell.value)
            data.append(col_data)
        return data

    def writeRowsInSheet(self,data):
        font = Font(name="Times New Roman",bold=False, size=8)
        side = Side(style='thin', color="000000")
        border = Border(left=side, top=side, right=side, bottom=side)
        alignment = Alignment(horizontal='left', vertical='bottom')
        if not isinstance(data,(tuple,list)):
            print("写入的数据不是列表或者元祖，请重试！")
            return
        for line in data:
            if not isinstance(line,(tuple,list)):
                print("写入的数据不是列表或者元祖，请重试！")
                return
            self.sheet.append(line)
        for rowNo,rowObj in enumerate(self.sheet.rows):
            for cell in rowObj:
                cell.font=font
                cell.border=border
                cell.alignment=alignment
                print("数据【%s】写入sheet【%s】完毕！"%(cell.value,self.sheet.title))


    def writeColumnInSheet(self,ColNo,data):
        if not isinstance(data,(tuple,list)):
            print("写入的数据不是列表或者元祖，请重试！")
            return
        if ColNo<1:
            print("输入的列号【%s】应该>=1"%ColNo)
            return
        for rowNo,value in enumerate(data):
            self.writeCellValue(rowNo+1,ColNo,value)
        print("数据写入sheet【%s】完毕！"%self.sheet.title)

    def createNewSheet(self,title=None,index=None):
        if title in self.getSheetNames():
            print("sheet【%s】已经存在！"%title)
            return
        self.wb.create_sheet(title=title,index=index)
        #self.wb.save(self.filePath)
        print("创建sheet【%s】完成！"%title)

    def getAllRows(self):
        rows=list(self.sheet.rows)
        return rows


    def getRow(self,rowNo):
        rows=[]
        for row in self.sheet.iter_rows():
            rows.append(row)
        if not isinstance(rowNo,int):
            print("行号输入错误，请输入整数！")
            return
        if rowNo>len(rows) or rowNo<1:
            print("输入的行号【%s】超出范围"%rowNo)
            return
        return rows[rowNo-1]

    def getRowValues(self,rowNo):
        values=[]
        if not self.getRow(rowNo):
            return
        for cell in self.getRow(rowNo):
            values.append(cell.value)
        return values

    def getColumn(self,colNo):
        cols=[]
        for col in self.sheet.iter_cols():
            cols.append(col)
        if not isinstance(colNo,int):
            print("列号输入错误，请输入整数！")
            return
        if colNo>len(cols) or colNo<1:
            print("输入的列号【%s】超出范围"%colNo)
            return
        return cols[colNo-1]

    def getColValues(self,colNo):
        values=[]
        if not self.getColumn(colNo):
            return
        for cell in self.getColumn(colNo):
            values.append(cell.value)
        return values

    def getCellValue(self,rowNo,colNo):
        if not (isinstance(rowNo,int) and isinstance(colNo,int)):
            print("输入的行号【%s】或者列号【%s】数字类型有误,请重新输入！"%(rowNo,colNo))
            return
        if rowNo>self.getMaxRowNo() or rowNo<1:
            print("输入的行号超出范围，应该1<= <=%s"%self.getMaxRowNo())
            return None
        if colNo>self.getMaxColumnNo() or colNo<1:
            print("输入的列号超出范围，应该1<= <=%s"%self.getMaxColumnNo())
            return None
        return self.sheet.cell(row=rowNo,column=colNo).value

    def writeCellValue(self,rowNo,colNo,value,color=None):
        if not (isinstance(rowNo,int) and isinstance(colNo,int)):
            print("输入的行号【%s】或者列号【%s】数字类型有误,请重新输入！"%(rowNo,colNo))
            return
        if rowNo<1:
            print("输入的行号超出范围，应该>=1")
            return None
        if colNo<1:
            print("输入的列号超出范围，应该1>=")
            return None
        if color:
            if "red" in color:
                self.sheet.cell(row=rowNo, column=colNo).font = Font(color=colors.RED)
            elif "green" in color:
                self.sheet.cell(row=rowNo, column=colNo).font = Font(color=colors.GREEN)
            else:
                self.sheet.cell(row=rowNo, column=colNo).font = Font(color=colors.BLACK)
        side= Side(style='thin', color="000000")
        self.sheet.cell(row=rowNo, column=colNo).border = Border(left=side, top=side, right=side, bottom=side)
        self.sheet.cell(row=rowNo,column=colNo,value=value)
        print("数据【%s】写入sheet【%s】完毕！" % (value, self.sheet.title))

    def writeTime(self,rowNo,colNo):
        currentTime=TimeUtil().get_chinesedatetime()
        print(currentTime)
        self.sheet.cell(row=rowNo,column=colNo,value=currentTime)

    def addRowStyle(self,rowObj):
        font = Font(name="黑体",bold=False, size=8)
        side = Side(style='thin', color="000000")
        border = Border(left=side, top=side, right=side, bottom=side)
        alignment = Alignment(horizontal='left', vertical='bottom')
        fill=PatternFill("solid",fgColor="FFBB66")
        for cell in rowObj:
            print(cell)
            cell.font=font
            cell.border=border
            cell.alignment=alignment
            cell.fill=fill
        print("行加入样式！")

    def addCellStyle(self,cellObj,color):
        if color=="red":
            cellObj.font= Font(name="Times New Roman",size=8,bold=True,color=colors.RED)
        elif color=="green":
            cellObj.font= Font(name="Times New Roman",size=8,bold=True,color=colors.GREEN)
        print("单元格加入样式！")

    def saveExcel(self):
        self.wb.save(self.filePath)
        print("excel【%s】保存成功！"%(os.path.basename(self.filePath)))





if __name__=="__main__":
    #excel_path=ExcelParser("d:\\pydelete\\0704")
    excelObj= ExcelParser(r"E:\我的坚果云\framework\practice\web_ui_po\TestData\126邮箱联系人.xlsx")
    # print(excelObj.getSheetNames())
    # print(excelObj.getSheetByIndex(1))
    # print(excelObj.getSheetByName("联系人2"))
    # print(excelObj.getSheetByName("126账号"))
    # print(excelObj.getMaxRowNo())
    # print(excelObj.getMaxColumnNo())
    # print(excelObj.getSheetAllCellValuesByRow())
    # print(excelObj.getSheetAllCellValuesByColumn())
    excelObj.createNewSheet("ssstttt")
    excelObj.getSheetByName("ssstttt")
    excelObj.writeRowsInSheet([("a","b","c"),(11,33,44)])
    # excelObj.writeColumnInSheet(11,["columntest","a","c"])
    rowObj=excelObj.getRow(1)
    # print(excelObj.getRowValues(1))
    print("ddd",excelObj.getColumn(1))
    # print(excelObj.getColValues(1))
    # print(excelObj.getCellValue(1.2,3))
    # print(excelObj.writeCellValue(4,16,"ssssssss","red"))
    # excelObj.writeTime(5,100)
    print("xx",excelObj.getAllRows())
    print(rowObj)
    excelObj.addRowStyle(rowObj)
    for cell in rowObj:
        excelObj.addCellStyle(cell,"red")
    excelObj.saveExcel()
